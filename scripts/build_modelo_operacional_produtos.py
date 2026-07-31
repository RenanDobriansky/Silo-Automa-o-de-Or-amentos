"""Gera a planilha operacional oficial do de/para de produtos."""

from __future__ import annotations

from pathlib import Path
import shutil
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


MODEL_OUTPUT_PATH = Path("data/apoio/modelo_operacional_tabela_produtos.xlsx")
OFFICIAL_OUTPUT_PATH = Path("data/apoio/Tabela de produtos.xlsx")
BACKUP_DIR = Path("data/apoio/_backup")
PRODUCTS_SOURCE_PATH = Path("data/saida/relatorios/produtos_unicos.xlsx")
BARCODE_SOURCE_PATH = Path("data/saida/relatorios/produtos_status_codigo_barras_syscomp.xlsx")

CADASTRO_HEADERS = [
    "Ativo",
    "Status Item",
    "Prioridade",
    "Data início",
    "Data fim",
    "Item",
    "COD. SILO",
    "Código de Barras",
    "DESCRIÇÃO",
    "CONVERSÃO",
    "Observação",
]


def _style_palette() -> dict[str, object]:
    thin = Side(style="thin", color="B7C9D6")
    return {
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "info_fill": PatternFill("solid", fgColor="D9EAF7"),
        "header_font": Font(color="FFFFFF", bold=True),
        "title_font": Font(bold=True, size=14),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(wrap_text=True, vertical="top"),
    }


def _build_cadastro_sheet(workbook: Workbook, styles: dict[str, object]) -> None:
    sheet = workbook.active
    sheet.title = "cadastro_produtos"
    rows = _load_operational_rows()

    sheet["A1"] = "Modelo Operacional De/Para de Produtos"
    sheet["A1"].font = styles["title_font"]
    sheet.merge_cells("A1:K1")

    sheet["A2"] = (
        "Use esta planilha para manter o codigo vigente dos itens. "
        "O sistema prioriza linhas ativas e com maior prioridade."
    )
    sheet["A2"].alignment = styles["wrap"]
    sheet.merge_cells("A2:K2")

    for column_index, value in enumerate(CADASTRO_HEADERS, start=1):
        cell = sheet.cell(row=4, column=column_index, value=value)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]

    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = styles["border"]
            cell.alignment = (
                styles["wrap"] if column_index not in (1, 2, 3, 4, 5, 7, 8) else styles["center"]
            )
            if row_index <= 8 and column_index <= 5:
                cell.fill = styles["info_fill"]

    last_data_row = max(5, 4 + len(rows))
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:K{last_data_row}"

    widths = {
        "A": 12,
        "B": 18,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 38,
        "G": 14,
        "H": 20,
        "I": 55,
        "J": 60,
        "K": 44,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    ativo_validation = DataValidation(type="list", formula1='"Sim,Nao"', allow_blank=True)
    sheet.add_data_validation(ativo_validation)
    ativo_validation.add(f"A5:A{max(last_data_row + 20, 200)}")

    status_validation = DataValidation(
        type="list",
        formula1='"Atendido,Nao atendido,Descontinuado"',
        allow_blank=True,
    )
    sheet.add_data_validation(status_validation)
    status_validation.add(f"B5:B{max(last_data_row + 20, 200)}")

    comments = {
        "A4": "Use Sim para a linha vigente e Nao para historico/inativo.",
        "B4": "Use Nao atendido para itens que a empresa nao vende e que devem bloquear a geracao automatica.",
        "C4": "Maior numero = maior prioridade entre linhas ativas do mesmo Item.",
        "D4": "Opcional. Se preenchida, a linha so vale a partir desta data.",
        "E4": "Opcional. Se preenchida, a linha deixa de valer apos esta data.",
        "H4": "Codigo de barras oficial atualmente conhecido no Syscomp. Se nao existir, fica em branco.",
        "J4": "Opcional. Regra de conversao comercial usada pelo projeto.",
    }
    for cell_ref, text in comments.items():
        sheet[cell_ref].comment = Comment(text, "Codex")


def _build_instruction_sheet(workbook: Workbook, styles: dict[str, object]) -> None:
    sheet = workbook.create_sheet("instrucoes")
    rows = [
        ["Passo", "O que fazer", "Exemplo", "Observação"],
        ["1", "Quando usar Ativo = Nao", "Codigo antigo do Doce de Goiaba", "Use quando a linha ficar apenas como historico ou nao puder mais ser usada automaticamente."],
        ["2", "Quando usar Status Item = Nao atendido", "Produto que aparece na cotacao mas a empresa nao vende", "Use quando o item deve continuar registrado na planilha, mas sem cadastro no sistema e sem atendimento comercial."],
        ["3", "Quando cadastrar novo COD. SILO", "Doce de Goiaba novo 999", "Cadastre uma nova linha quando o item passar a existir no sistema ou trocar para um novo codigo interno."],
        ["4", "Como trocar um item de um codigo antigo para um novo", "Linha antiga Ativo = Nao e linha nova Ativo = Sim", "Nao apague a linha antiga. Desative a antiga e crie ou mantenha a nova como vigente."],
        ["5", "Quando apenas deixar o item fora do atendimento", "Abacaxi em Calda sem produto no ERP", "Deixe Status Item = Nao atendido e mantenha COD. SILO, DESCRIÇÃO e Codigo de Barras em branco."],
        ["6", "Se houver mais de uma linha ativa para o mesmo Item", "Prioridade 10 vence prioridade 1", "O sistema escolhe a linha com maior prioridade entre as linhas ativas."],
        ["7", "Atualizacao de Codigo de Barras", "7891234567890", "Atualize quando o ERP passar a ter codigo oficial. Se ainda nao existir, deixe em branco."],
        ["8", "Uso opcional de Data início e Data fim", "01/08/2026", "Preencha apenas se quiser controlar vigencia por periodo."],
        ["9", "Nao altere o nome das colunas", "Item / COD. SILO / DESCRIÇÃO / Status Item", "O sistema depende desses nomes para ler a planilha."],
    ]

    sheet["A1"] = "Como usar a planilha"
    sheet["A1"].font = styles["title_font"]
    sheet.merge_cells("A1:D1")

    for row_index, values in enumerate(rows, start=3):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = styles["border"]
            cell.alignment = styles["wrap"]
            if row_index == 3:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = styles["center"]

    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 44
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 36


def _build_notes_sheet(workbook: Workbook, styles: dict[str, object]) -> None:
    sheet = workbook.create_sheet("observacoes_rapidas")
    rows = [
        ["Campo", "Quando preencher", "Como a equipe deve usar"],
        ["Item", "Sempre", "Escreva como o item costuma aparecer na cotacao ou na OC."],
        ["Status Item", "Sempre que o item nao for atendido ou estiver descontinuado", "Use Atendido para item normal. Use Nao atendido para item fora do portifolio. Use Descontinuado quando o item deixou de ser vendido."],
        ["COD. SILO", "Quando o item existir no sistema", "Preencha com o codigo interno do ERP. Se o item nao for atendido, pode deixar em branco."],
        ["Código de Barras", "Quando houver codigo oficial no ERP", "Atualize quando o cadastro do Syscomp passar a ter codigo de barras. Se ainda nao existir, deixe em branco."],
        ["DESCRIÇÃO", "Quando o item existir no sistema", "Use a descricao oficial do produto no ERP."],
        ["CONVERSÃO", "Quando o item precisar regra comercial", "Use para caixa fechada, pacote, arredondamento ou outras conversoes da operacao."],
        ["Ativo", "Quando houver linha vigente ou historica", "Use Sim para a linha atual. Use Nao para codigo antigo, historico ou linha que nao deve mais ser usada."],
        ["Prioridade", "Quando existir mais de uma linha possivel para o mesmo Item", "Maior numero vence entre linhas ativas."],
        ["Data início / Data fim", "Quando quiser controlar vigencia por periodo", "Opcional. Use se uma regra ou codigo so valer em determinado periodo."],
        ["Observação", "Quando precisar orientar a equipe", "Anote motivo da troca, regra comercial ou qualquer detalhe util para quem vai manter a planilha."],
        ["Resumo rapido", "Regra geral", "Se o item existe no sistema: preencha codigo, descricao e, se houver, codigo de barras. Se nao existe no sistema: marque Status Item = Nao atendido."],
    ]

    sheet["A1"] = "Guia rapido dos campos da planilha"
    sheet["A1"].font = styles["title_font"]
    sheet.merge_cells("A1:C1")

    sheet["A2"] = "Esta aba serve como referencia rapida para quem for manter a planilha no dia a dia."
    sheet["A2"].alignment = styles["wrap"]
    sheet.merge_cells("A2:C2")

    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = styles["border"]
            cell.alignment = styles["wrap"]
            if row_index == 4:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = styles["center"]

    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 62


def _load_operational_rows() -> list[list[object]]:
    """Converte a tabela tratada atual em linhas prontas para a planilha operacional."""

    if not PRODUCTS_SOURCE_PATH.exists():
        return [[
            "Sim",
            "Atendido",
            1,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Fonte produtos_unicos.xlsx nao encontrada",
        ]]

    dataframe = pd.read_excel(PRODUCTS_SOURCE_PATH).fillna("")
    barcode_lookup = _load_barcode_lookup()
    rows: list[list[object]] = []

    for _, row in dataframe.iterrows():
        codigo = _normalize_excel_value(_read_column(row, "COD. SILO"))
        descricao = str(_read_column(row, "DESCRIÇÃO", "DESCRICAO") or "").strip()
        codigo_barras = barcode_lookup.get(codigo, "")
        status_item = _infer_item_status(codigo, descricao)
        observacao = _append_notes(_build_observation(row), codigo, descricao, codigo_barras)
        rows.append(
            [
                "Sim",
                status_item,
                1,
                "",
                "",
                str(_read_column(row, "Item") or "").strip(),
                codigo,
                codigo_barras,
                descricao,
                str(_read_column(row, "CONVERSÃO", "CONVERSAO") or "").strip(),
                observacao,
            ]
        )

    rows.append(["", "", "", "", "", "", "", "", "", "", "Novos itens podem ser adicionados abaixo desta linha."])
    return rows


def _load_barcode_lookup() -> dict[str, str]:
    """Carrega o codigo de barras oficial conhecido para cada COD. SILO."""

    if not BARCODE_SOURCE_PATH.exists():
        return {}

    dataframe = pd.read_excel(BARCODE_SOURCE_PATH).fillna("")
    lookup: dict[str, str] = {}
    for _, row in dataframe.iterrows():
        codigo_silo = _normalize_excel_value(_read_column(row, "codigo_silo", "COD. SILO"))
        codigo_barras = _normalize_excel_value(
            _read_column(row, "codigo_barras_oficial", "Código de Barras", "Codigo de Barras")
        )
        if codigo_silo:
            lookup[codigo_silo] = codigo_barras
    return lookup


def _normalize_excel_value(value: object) -> str:
    """Converte valores numericos do Excel para texto sem sufixo .0."""

    if value in ("", None):
        return ""

    text = str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    if text.endswith(".00"):
        return text[:-3]
    return text


def _build_observation(row: pd.Series) -> str:
    """Monta uma observacao inicial com base no estado atual do cadastro."""

    item = str(_read_column(row, "Item") or "").strip()
    codigo = _normalize_excel_value(_read_column(row, "COD. SILO"))
    descricao = str(_read_column(row, "DESCRIÇÃO", "DESCRICAO") or "").strip()
    criterio = str(_read_column(row, "criterio_escolha") or "").strip()

    notes: list[str] = []
    if not codigo:
        notes.append("Sem COD. SILO definido")
    if not descricao:
        notes.append("Sem DESCRIÇÃO definida")
    if criterio and criterio != "item_unico_ou_sem_conflito":
        notes.append(criterio)
    if not notes and item:
        notes.append("Carga inicial automatica do cadastro atual")
    return " | ".join(notes)


def _append_notes(observacao: str, codigo_silo: str, descricao: str, codigo_barras: str) -> str:
    """Complementa a observacao com o status do item e do codigo de barras."""

    notes = [part.strip() for part in str(observacao or "").split("|") if part.strip()]
    if not codigo_silo or not descricao:
        notes.append("Item nao atendido pela empresa")
    if codigo_silo and not codigo_barras:
        notes.append("Sem codigo de barras oficial")
    return " | ".join(dict.fromkeys(notes))


def _infer_item_status(codigo_silo: str, descricao: str) -> str:
    """Define o status inicial do item para a planilha operacional."""

    if not codigo_silo or not descricao:
        return "Nao atendido"
    return "Atendido"


def _read_column(row: pd.Series, *candidates: str) -> object:
    """Le uma coluna aceitando variacoes de acento e capitalizacao."""

    lookup = {_normalize_column_name(column): value for column, value in row.items()}
    for candidate in candidates:
        normalized_candidate = _normalize_column_name(candidate)
        if normalized_candidate in lookup:
            return lookup[normalized_candidate]
    return ""


def _normalize_column_name(value: object) -> str:
    """Normaliza nomes de colunas para comparacao resiliente."""

    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.casefold()


def _backup_official_file() -> Path | None:
    """Cria backup do arquivo oficial atual antes da substituicao."""

    if not OFFICIAL_OUTPUT_PATH.exists():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"Tabela de produtos_backup_{pd.Timestamp.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(OFFICIAL_OUTPUT_PATH, backup_path)
    return backup_path


def main() -> None:
    styles = _style_palette()
    workbook = Workbook()
    _build_cadastro_sheet(workbook, styles)
    _build_instruction_sheet(workbook, styles)
    _build_notes_sheet(workbook, styles)

    MODEL_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    backup_path = _backup_official_file()
    if MODEL_OUTPUT_PATH.exists():
        MODEL_OUTPUT_PATH.unlink()
    if OFFICIAL_OUTPUT_PATH.exists():
        OFFICIAL_OUTPUT_PATH.unlink()

    workbook.save(MODEL_OUTPUT_PATH)
    workbook.save(OFFICIAL_OUTPUT_PATH)
    print(MODEL_OUTPUT_PATH.resolve())
    print(OFFICIAL_OUTPUT_PATH.resolve())
    if backup_path is not None:
        print(backup_path.resolve())


if __name__ == "__main__":
    main()
